import pandas as pd
import json
import openpyxl
import os
import tkinter as tk
from tkinter import filedialog

# --- 參數設定 ---
USER_NAME = "李威漢"

def generate_final_json():
    # 建立選檔視窗
    root = tk.Tk()
    root.withdraw()  # 隱藏主視窗

    print("請選擇您的 App JSON 原始檔案 (例如 records_cleaned.json)...")
    json_path = filedialog.askopenfilename(
        title="選擇 App JSON 檔案",
        filetypes=[("JSON files", "*.json")]
    )
    
    print("請選擇公司出勤紀錄 Excel (.xlsx)...")
    excel_path = filedialog.askopenfilename(
        title="選擇公司 Excel 檔案",
        filetypes=[("Excel files", "*.xlsx")]
    )

    if not json_path or not excel_path:
        print("未選擇檔案，程式終止。")
        return

    # 1. 載入 JSON 出差參考 (補足週末出差用)
    with open(json_path, 'r', encoding='utf-8') as f:
        app_data = json.load(f)
    travel_map = {}
    for entry in app_data:
        try:
            d = pd.to_datetime(entry.get('date')).date()
            travel_map[d] = entry.get('travelCountry', "")
        except: continue

    # 2. 使用 openpyxl 讀取 Excel (支援顏色偵測)
    print(f"正在讀取 Excel: {os.path.basename(excel_path)} ...")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    all_output = []

    # 判斷是否為 2025 年份 (決定是否啟用藍底加班費邏輯)
    is_2025 = "2025" in os.path.basename(excel_path)

    # 遍歷 1-12 月分頁
    for m in range(1, 13):
        s_attn = f"TP出勤{m}月份"
        s_leave = f"TP{m}月份請假"
        s_ot = f"TP{m}月份加班"

        if s_attn not in wb.sheetnames: continue
        
        ws_a = wb[s_attn]
        ws_l = wb[s_leave] if s_leave in wb.sheetnames else None
        ws_o = wb[s_ot] if s_ot in wb.sheetnames else None

        # 定位「李威漢」所在的列 (Row)
        target_row = None
        for r in range(1, ws_a.max_row + 1):
            if str(ws_a.cell(r, 2).value).strip() == USER_NAME:
                target_row = r
                break
        
        if not target_row: continue

        # 定位日期標題列並遍歷每一天 (Column)
        for c in range(3, ws_a.max_column + 1):
            # 假設日期在第 1 或 2 列
            d_val = ws_a.cell(1, c).value or ws_a.cell(2, c).value
            try:
                date_obj = pd.to_datetime(d_val).date()
            except: continue

            # 初始結構 (遵循薪資計算日期依照公司)
            entry = {
                "date": date_obj.isoformat(),
                "otHours": 0.0,
                "leaveAmount": 0.0,
                "travelCountry": travel_map.get(date_obj, ""), # 週末依賴 JSON 補全
                "Remark": None,
                "otType": "pay", 
                "isLeave": False,
                "recordType": "attendance"
            }

            # A. 出勤與 ◎ 處理
            attn_val = str(ws_a.cell(target_row, c).value or "")
            if "◎" in attn_val:
                entry["Remark"] = "部門內部補休"

            # B. 請假處理 (單位 0.5 -> 1hr)
            if ws_l:
                l_val = ws_l.cell(target_row, c).value
                if l_val and isinstance(l_val, (int, float)) and l_val > 0:
                    entry["leaveAmount"] = l_val / 2.0
                    entry["isLeave"] = True

            # C. 加班處理與顏色辨識
            if ws_o:
                o_val = ws_o.cell(target_row, c).value
                if o_val and isinstance(o_val, (int, float)) and o_val > 0:
                    entry["otHours"] = o_val / 2.0
                    
                    # 2025 特有藍底邏輯：無色或白色則視為部門內部補休
                    if is_2025:
                        fill = ws_o.cell(target_row, c).fill
                        color_idx = fill.start_color.index if fill else "00000000"
                        # 如果是無底色 (00000000) 或白色 (FFFFFFFF)
                        if color_idx in ["00000000", "FFFFFFFF", "00FFFFFF", 64]:
                            entry["otType"] = "internal" 
                            entry["Remark"] = (entry["Remark"] + "; " if entry["Remark"] else "") + "內部補休(無藍底)"
                        else:
                            entry["otType"] = "pay" # 有藍底才算加班費

            all_output.append(entry)

    # 選擇存檔位置
    save_path = filedialog.asksaveasfilename(
        defaultextension=".json",
        initialfile=f"app_synced_{os.path.basename(excel_path).replace('.xlsx', '')}.json",
        title="儲存整合後的 JSON"
    )

    if save_path:
        with open(save_path, 'w', encoding='utf-8') as f:
            json.dump(all_output, f, ensure_ascii=False, indent=2)
        print(f"\n轉換成功！產出檔案：{save_path}")

if __name__ == "__main__":
    generate_final_json()